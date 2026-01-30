#!/usr/bin/env python3
import openpyxl
import json

# Load the workbook
wb = openpyxl.load_workbook('Track 201 Module 3 Lesson 2- Analytics Matrix.xlsx')
ws = wb.active

# Extract MLB data (rows 4-11)
mlb_players = []
for row in range(4, 12):
    player = {
        "name": ws.cell(row=row, column=2).value,
        "obp": ws.cell(row=row, column=3).value,
        "slg": ws.cell(row=row, column=4).value,
        "ops": ws.cell(row=row, column=5).value,
        "war": ws.cell(row=row, column=6).value,
        "wrc_plus": ws.cell(row=row, column=7).value,
        "salary": ws.cell(row=row, column=8).value
    }
    mlb_players.append(player)

# Extract NBA data (rows 16-23)
nba_players = []
for row in range(16, 24):
    player = {
        "name": ws.cell(row=row, column=2).value,
        "ts_percent": ws.cell(row=row, column=3).value,
        "usage_percent": ws.cell(row=row, column=4).value,
        "bpm": ws.cell(row=row, column=5).value,
        "ortg": ws.cell(row=row, column=6).value,
        "drtg": ws.cell(row=row, column=7).value,
        "salary": ws.cell(row=row, column=8).value
    }
    nba_players.append(player)

# Extract NFL data (rows 28-35)
nfl_players = []
for row in range(28, 36):
    player = {
        "name": ws.cell(row=row, column=2).value,
        "epa_per_play": ws.cell(row=row, column=3).value,
        "cpoe": ws.cell(row=row, column=4).value,
        "success_rate": ws.cell(row=row, column=5).value,
        "dvoa": ws.cell(row=row, column=6).value,
        "age": ws.cell(row=row, column=7).value,
        "salary": ws.cell(row=row, column=8).value
    }
    nfl_players.append(player)

# Create scenarios structure
scenarios = [
    {
        "id": 1,
        "sport": "MLB",
        "title": "Replace a Star Hitter",
        "description": "Your star player just left in free agency. Build a replacement strategy using 2-4 players within budget.",
        "budget": 40,
        "minPlayers": 2,
        "maxPlayers": 4,
        "passingCriteria": {
            "obp_avg": {"threshold": 0.33, "operator": ">=", "label": "OBP Average"},
            "wrc_plus_avg": {"threshold": 105, "operator": ">=", "label": "wRC+ Average"},
            "war_efficiency": {"threshold": 0.12, "operator": ">=", "label": "WAR per $M"}
        },
        "statLabels": {
            "obp": "OBP",
            "slg": "SLG",
            "ops": "OPS",
            "war": "WAR",
            "wrc_plus": "wRC+",
            "salary": "Salary"
        },
        "players": mlb_players
    },
    {
        "id": 2,
        "sport": "NBA",
        "title": "Build an Efficient Playoff Lineup",
        "description": "Construct a balanced playoff rotation with 3-5 players who can score efficiently and play defense.",
        "budget": 75,
        "minPlayers": 3,
        "maxPlayers": 5,
        "passingCriteria": {
            "ts_percent_avg": {"threshold": 0.57, "operator": ">=", "label": "TS% Average"},
            "bpm_avg": {"threshold": 1.5, "operator": ">=", "label": "BPM Average"},
            "ortg_avg": {"threshold": 113, "operator": ">=", "label": "ORtg Average"},
            "drtg_avg": {"threshold": 115, "operator": "<=", "label": "DRtg Average"}
        },
        "statLabels": {
            "ts_percent": "TS%",
            "usage_percent": "Usage%",
            "bpm": "BPM",
            "ortg": "ORtg",
            "drtg": "DRtg",
            "salary": "Salary"
        },
        "players": nba_players
    },
    {
        "id": 3,
        "sport": "NFL",
        "title": "Fix a Broken Offense",
        "description": "Your offense ranked 28th last season. Pick 3-5 players to turn things around within budget.",
        "budget": 80,
        "minPlayers": 3,
        "maxPlayers": 5,
        "passingCriteria": {
            "epa_per_play_avg": {"threshold": 0.05, "operator": ">=", "label": "EPA/play Average"},
            "cpoe_avg": {"threshold": 1.0, "operator": ">=", "label": "CPOE Average"},
            "success_rate_avg": {"threshold": 48, "operator": ">=", "label": "Success Rate Average"}
        },
        "statLabels": {
            "epa_per_play": "EPA/play",
            "cpoe": "CPOE",
            "success_rate": "Success Rate",
            "dvoa": "DVOA",
            "age": "Age",
            "salary": "Salary"
        },
        "players": nfl_players
    }
]

# Output as JSON
print(json.dumps(scenarios, indent=2))
